#!/usr/bin/env python3
"""Generate the expense report workbook and combined PDF."""

from __future__ import annotations

import argparse
import json
import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape as xml_escape

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from pypdf import PdfReader, PdfWriter
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ModuleNotFoundError as exc:
    print(
        f"Missing Python package: {exc.name}\n"
        "Install or update dependencies with:\n"
        "  python -m pip install --upgrade openpyxl pypdf reportlab",
        file=sys.stderr,
    )
    raise SystemExit(1) from exc


DEFAULT_WORKBOOK = "Expense Report Guide.xlsx"
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def money(value: Any) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def money_text(value: Decimal) -> str:
    text = f"{value:.2f}"
    return text.rstrip("0").rstrip(".")


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def xlsx_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value)
    if text.startswith(FORMULA_PREFIXES):
        return "'" + text
    return text


def pdf_text(value: Any) -> str:
    return xml_escape(as_text(value))


def clean_filename_part(value: str) -> str:
    cleaned = "".join(char for char in value.strip() if char not in '/\\:*?"<>|')
    cleaned = " ".join(cleaned.split())
    if not cleaned:
        raise ValueError("Filename part is empty after sanitization.")
    return cleaned


def clean_output_filename(value: str, expected_suffix: str) -> str:
    name = clean_filename_part(Path(value).name)
    if Path(name).suffix.lower() != expected_suffix:
        raise ValueError(f"Output filename must end in {expected_suffix}: {value}")
    return name


def resolve_pdf_name(manifest: dict[str, Any], user_name: str | None, explicit_pdf_name: str | None) -> str:
    if explicit_pdf_name:
        return clean_output_filename(explicit_pdf_name, ".pdf")
    report_user_name = as_text(user_name or manifest.get("report_user_name") or manifest.get("user_name")).strip()
    if not report_user_name:
        raise ValueError("Provide report_user_name in the manifest or pass --user-name for the default PDF filename.")
    return f"expense report {clean_filename_part(report_user_name)}.pdf"


def resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = base_dir / path
    return path


def load_manifest(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        manifest = json.load(handle)
    if not isinstance(manifest, dict):
        raise ValueError("Manifest must be a JSON object.")
    return manifest


def normalize_expenses(manifest: dict[str, Any], manifest_dir: Path) -> list[dict[str, Any]]:
    raw_expenses = manifest.get("expenses", manifest.get("receipts"))
    if not isinstance(raw_expenses, list) or not raw_expenses:
        raise ValueError("Manifest must contain a non-empty 'expenses' list.")

    expenses: list[dict[str, Any]] = []
    required = ["source_pdf", "expense_type", "amount_chf", "reason", "project"]
    for index, raw in enumerate(raw_expenses, start=1):
        if not isinstance(raw, dict):
            raise ValueError(f"Expense row {index} must be an object.")
        missing = [key for key in required if raw.get(key) in (None, "")]
        if missing:
            raise ValueError(f"Expense row {index} is missing: {', '.join(missing)}.")

        row = dict(raw)
        row["amount_chf"] = money(row["amount_chf"])
        row["currency"] = as_text(row.get("currency") or "CHF")
        original_currency = as_text(row.get("original_currency")).upper()
        if original_currency and original_currency != "CHF" and not as_text(row.get("rate_note")).strip():
            raise ValueError(f"Expense row {index} uses {original_currency}; add a trusted-source rate_note.")
        if row.get("source_pdf"):
            source_pdf = resolve_path(as_text(row["source_pdf"]), manifest_dir)
            if not source_pdf.exists():
                raise FileNotFoundError(f"Missing receipt PDF for row {index}: {source_pdf}")
            row["source_pdf"] = source_pdf
        expenses.append(row)
    return expenses


def normalize_car_trips(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    raw_trips = manifest.get("car_trips") or []
    if not isinstance(raw_trips, list):
        raise ValueError("'car_trips' must be a list when provided.")
    trips: list[dict[str, Any]] = []
    for index, raw in enumerate(raw_trips, start=1):
        if not isinstance(raw, dict):
            raise ValueError(f"Car trip row {index} must be an object.")
        row = dict(raw)
        row["km_amount"] = money(row.get("km_amount", 0))
        trips.append(row)
    return trips


def write_workbook(path: Path, expenses: list[dict[str, Any]], car_trips: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Sheet1"

    ws.merge_cells("A1:E1")
    ws.merge_cells("H1:N1")
    ws["A1"] = "Travel Expenses Private Car"
    ws["H1"] = "Expenses with Receipt"

    car_headers = ["Date", "km amount", "Start", "End", "Resaon", "Project"]
    receipt_headers = [
        "Date",
        "Nr. (Receipt ID)",
        "Expense Type",
        "Amount in CHF",
        "Currency ",
        "Rate",
        "Reason",
        "Project",
    ]
    for col, label in enumerate(car_headers, start=1):
        ws.cell(row=2, column=col, value=label)
    for col, label in enumerate(receipt_headers, start=8):
        ws.cell(row=2, column=col, value=label)

    for offset, trip in enumerate(car_trips, start=3):
        values = [
            xlsx_text(trip.get("date")),
            float(trip["km_amount"]),
            xlsx_text(trip.get("start")),
            xlsx_text(trip.get("end")),
            xlsx_text(trip.get("reason")),
            xlsx_text(trip.get("project")),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=offset, column=col, value=value)

    for offset, expense in enumerate(expenses, start=3):
        values = [
            xlsx_text(expense.get("date")),
            xlsx_text(expense.get("receipt_id")),
            xlsx_text(expense["expense_type"]),
            float(expense["amount_chf"]),
            xlsx_text(expense["currency"]),
            xlsx_text(expense.get("rate_note")),
            xlsx_text(expense["reason"]),
            xlsx_text(expense["project"]),
        ]
        for col, value in enumerate(values, start=8):
            ws.cell(row=offset, column=col, value=value)

    first_row = 3
    car_end_row = max(first_row, first_row + len(car_trips) - 1)
    expense_end_row = max(first_row, first_row + len(expenses) - 1)
    total_row = max(18, expense_end_row + 2, car_end_row + 2)
    grand_total_row = total_row + 1

    ws.cell(row=total_row, column=2, value=f"=SUM(B3:B{car_end_row})*0.5")
    if expenses:
        ws.cell(row=total_row, column=8, value="Total:")
        ws.cell(row=total_row, column=9, value=f"=SUM(K3:K{expense_end_row})")
    else:
        ws.cell(row=total_row, column=8, value="Total:")
        ws.cell(row=total_row, column=9, value=0)
    ws.cell(row=grand_total_row, column=8, value="Grand Total")
    ws.cell(row=grand_total_row, column=9, value=f"=I{total_row}+B{total_row}")

    apply_workbook_style(ws, total_row, grand_total_row)
    workbook.save(path)


def apply_workbook_style(ws: Any, total_row: int, grand_total_row: int) -> None:
    widths = {
        "A": 9.83,
        "B": 16.66,
        "C": 12.5,
        "D": 12.66,
        "E": 24.5,
        "F": 14.83,
        "G": 1.83,
        "H": 9.83,
        "I": 10.83,
        "J": 10,
        "K": 11.33,
        "L": 7.16,
        "M": 58,
        "N": 42,
        "O": 24,
        "P": 10,
        "Q": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAD3")

    for cell in ("A1", "H1"):
        ws[cell].font = Font(name="Calibri", size=20, bold=True)

    for row in ws.iter_rows(min_row=2, max_row=grand_total_row, min_col=1, max_col=15):
        for cell in row:
            if cell.row == 2:
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.fill = header_fill
            else:
                cell.font = Font(name="Calibri", size=10)
            if 1 <= cell.column <= 6 or 8 <= cell.column <= 15:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row in (total_row, grand_total_row):
        for column in range(1, 16):
            ws.cell(row=row, column=column).font = Font(name="Calibri", size=10, bold=True)

    for row in range(3, grand_total_row + 1):
        ws.cell(row=row, column=2).number_format = "0.00"
        ws.cell(row=row, column=9).number_format = "0.00"
        ws.cell(row=row, column=11).number_format = "0.00"
        ws.row_dimensions[row].height = 28

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"H1:Q{grand_total_row}"

    for column in range(1, 18):
        ws.cell(row=1, column=column).alignment = Alignment(vertical="center")


def build_summary_pdf(path: Path, expenses: list[dict[str, Any]], total: Decimal) -> int:
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        str(path),
        pagesize=page_size,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExpenseTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        "ExpenseCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.5,
        alignment=TA_LEFT,
    )
    right_style = ParagraphStyle(
        "ExpenseRight",
        parent=cell_style,
        alignment=TA_RIGHT,
    )

    headers = [
        "Date",
        "Nr. (Receipt ID)",
        "Expense Type",
        "Amount in CHF",
        "Currency",
        "Reason",
        "Project",
    ]
    data: list[list[Any]] = [[Paragraph(header, cell_style) for header in headers]]
    for expense in expenses:
        data.append(
            [
                Paragraph(pdf_text(expense.get("date")), cell_style),
                Paragraph(pdf_text(expense.get("receipt_id")), cell_style),
                Paragraph(pdf_text(expense["expense_type"]), cell_style),
                Paragraph(money_text(expense["amount_chf"]), right_style),
                Paragraph(pdf_text(expense.get("currency") or "CHF"), cell_style),
                Paragraph(pdf_text(expense["reason"]), cell_style),
                Paragraph(pdf_text(expense["project"]), cell_style),
            ]
        )

    table = Table(
        data,
        colWidths=[24 * mm, 30 * mm, 24 * mm, 28 * mm, 22 * mm, 95 * mm, 50 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAD3")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    totals = Table(
        [
            [Paragraph("Total:", cell_style), Paragraph(money_text(total), right_style)],
            [Paragraph("Grand Total", cell_style), Paragraph(money_text(total), right_style)],
        ],
        colWidths=[34 * mm, 28 * mm],
        hAlign="LEFT",
    )
    totals.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    doc.build([Paragraph("Expenses with Receipt", title_style), table, Spacer(1, 6 * mm), totals])
    reader = PdfReader(str(path))
    return len(reader.pages)


def merge_pdf(summary_pdf: Path, receipt_pdfs: list[Path], output_pdf: Path) -> int:
    writer = PdfWriter()
    for source in [summary_pdf, *receipt_pdfs]:
        reader = PdfReader(str(source))
        for page in reader.pages:
            writer.add_page(page)
    with output_pdf.open("wb") as handle:
        writer.write(handle)
    return len(writer.pages)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate expense report XLSX and PDF.")
    parser.add_argument("--manifest", required=True, help="Path to the reviewed JSON manifest.")
    parser.add_argument("--out-dir", required=True, help="Directory for final outputs.")
    parser.add_argument("--workbook-name", default=DEFAULT_WORKBOOK)
    parser.add_argument("--pdf-name", help="Override the default PDF name.")
    parser.add_argument("--user-name", help="Report user name for the default PDF filename.")
    args = parser.parse_args()

    manifest_path = Path(args.manifest).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest = load_manifest(manifest_path)
    expenses = normalize_expenses(manifest, manifest_path.parent)
    car_trips = normalize_car_trips(manifest)

    workbook_path = out_dir / clean_output_filename(args.workbook_name, ".xlsx")
    pdf_name = resolve_pdf_name(manifest, args.user_name, args.pdf_name)
    output_pdf = out_dir / pdf_name
    summary_pdf = out_dir / f".{output_pdf.stem}-summary.pdf"

    write_workbook(workbook_path, expenses, car_trips)
    receipt_total = sum((expense["amount_chf"] for expense in expenses), Decimal("0.00"))
    car_total = sum((trip["km_amount"] for trip in car_trips), Decimal("0.00")) * Decimal("0.50")
    grand_total = (receipt_total + car_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    summary_pages = build_summary_pdf(summary_pdf, expenses, grand_total)

    receipt_pdfs = [
        Path(expense["source_pdf"])
        for expense in expenses
        if expense.get("source_pdf") and Path(expense["source_pdf"]).resolve() != output_pdf
    ]
    page_count = merge_pdf(summary_pdf, receipt_pdfs, output_pdf)
    try:
        os.remove(summary_pdf)
    except FileNotFoundError:
        pass

    print(json.dumps(
        {
            "workbook": str(workbook_path),
            "pdf": str(output_pdf),
            "report_user_name": as_text(args.user_name or manifest.get("report_user_name") or manifest.get("user_name")),
            "expense_rows": len(expenses),
            "receipt_pdfs": len(receipt_pdfs),
            "summary_pages": summary_pages,
            "pdf_pages": page_count,
            "grand_total_chf": money_text(grand_total),
        },
        indent=2,
    ))


if __name__ == "__main__":
    try:
        main()
    except (ValueError, FileNotFoundError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
